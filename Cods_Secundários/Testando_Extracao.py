# -*- coding: utf-8 -*-
#python -m pip install --upgrade certifi --trusted-host pypi.org --trusted-host files.pythonhosted.org ##### Para Atualizar Certificados Sem Usar o Pip Install Afetado
# Novo Código Reports

import pandas as pd
from openpyxl import load_workbook
from brfinance.connector import CVMHttpClientConnector
from brfinance.http_client import CVMHttpClient
from brfinance import CVMAsyncBackend
from datetime import datetime, date
import datetime
import openpyxl
import tkinter as tk
from tkinter import filedialog
import numpy as np
import os
import urllib2 # Importar a biblioteca urllib2
from bs4 import BeautifulSoup # Importar a biblioteca BeautifulSoup

# Criar uma instância da classe CVMAsyncBackend
cvm_httpclient = CVMAsyncBackend()
cvm = CVMHttpClient(CVMHttpClientConnector)

# Dict de códigos CVM para todas as empresas
cvm_codes = cvm_httpclient.get_cvm_codes()
# Busca Notícias e Fatos Relevante, ITRs
search_result = cvm_httpclient.get_consulta_externa_cvm_results(start_date=date(2023, 1, 1),
                                            end_date=date.today(),
                                            cod_cvm=['18724'],                      #last_ref_date = False # Se "True" retorna apenas o último report no intervalo de datas
                                            category=["EST_-1",                     #  'EST_-1': TODOS os Documentos Estruturados
                                                    "EST_3",                        #  'EST_1': 'FCA - Formulário Cadastral'
                                                    "IPE_4_-1_-1",                  #  'EST_2': 'FRE - Formulário de Referência'
                                                    "IPE_-1_58_-1"]                 #  'EST_3': 'ITR - Informações Trimestrais' 
                                                )                                   #  'EST_4': 'DFP - Demonstrações Financeiras Padronizadas'
                                                                                    #  'EST_11': 'Informe do Código de Governança'    
                                                                                    #  'EST_13': 'IAN - Informações Anuais'   
                                                                                    #  'IPE_70_-1_-1': 'Estatuto Social|TODOS|   TODAS'
                                                                                    #  'IPE_4_-1_-1': 'Fato Relevante|TODOS|   TODAS'   
                                                                                    #  'IPE_44_-1_-1': 'Acordo de Acionistas|TODOS|   TODAS'
                                                                                    #  'IPE_1_-1_-1': 'Assembleia|TODOS|   TODAS'
                                                                                    #  'IPE_1_98_-1': 'Assembleia|AGCRA|   TODAS'
                                                                                    #  'IPE_1_97_-1': 'Assembleia|AGCRI|   TODAS'
                                                                                    #  'IPE_1_5_-1': 'Assembleia|AGDEB|   TODAS'
                                                                                    #  'IPE_1_1_-1': 'Assembleia|AGE|   TODAS'
                                                                                    #  'IPE_91_113_-1': 'Carta Anual de Governança Corporativa|Carta Anual de Governança Corporativa (art. 8º  VIII da Lei 13.303/16|   TODAS'
                                                                                    #  'IPE_6_-1_-1': 'Comunicado ao Mercado|TODOS|   TODAS'
                                                                                    #  'IPE_6_53_-1': 'Comunicado ao Mercado|Outros Comunicados Não Considerados Fatos Relevantes|   TODAS'
                                                                                    #  'IPE_7_-1_-1': 'Dados Econômico-Financeiros|TODOS|   TODAS'
                                                                                    #  'IPE_7_56_-1': 'Dados Econômico-Financeiros|Balanço Social|   TODAS'
                                                                                    #  'IPE_7_37_-1': 'Dados Econômico-Financeiros|Demonstrações Financeiras Anuais Completas|   TODAS'
                                                                                    #  'IPE_7_46_-1': 'Dados Econômico-Financeiros|Demonstrações Financeiras Intermediárias|   TODAS'
                                                                                    #  'IPE_7_85_-1': 'Dados Econômico-Financeiros|Plano de Investimento|   TODAS'
                                                                                    #  'IPE_101_-1_-1': 'DF - Patrimônio separado|TODOS|   TODAS'
                                                                                    #  'IPE_78_-1_-1': 'Documentos de Oferta de Distribuição Pública|TODOS|   TODAS'
                                                                                    #  'IPE_54_-1_-1': 'Escrituras e aditamentos de debêntures|TODOS|   TODAS'
                                                                                    #  'IPE_58_-1_-1': 'Informações de Companhias em Recuperação Judicial ou Extrajudicial|TODOS|   TODAS'
                                                                                    #  'IPE_100_-1_-1': 'Informe Mensal de CRA|TODOS|   TODAS'
                                                                                    #  'IPE_83_-1_-1': 'Política de Dividendos|TODOS|   TODAS'
                                                                                    #  'IPE_76_-1_-1': 'Regimento Interno do Comitê de Auditoria Estatutário|TODOS|   TODAS'
                                                                                    #  'IPE_99_117_-1': 'Regulamentos da B3|Processo de enforcement|   TODAS'
                                                                                    #  'IPE_-1_37_-1': 'Tipo: Demonstrações Financeiras Anuais Completas'
                                                                                    #  'IPE_-1_58_-1': 'Tipo: Relatório Anual'
                                                                                    #  'IPE_-1_43_-1': 'Tipo: Relatório de Análise Gerencial'
                                                                                    #  'IPE_-1_91_-1': 'Tipo: Aviso ao Mercado'
                                                                                    #  'IPE_-1_138_-1': 'Tipo: Composição de carteira'
                                                                                    #  'IPE_-1_68_-1': 'Tipo: Causas e circunstâncias da falência'
                                                                                    #  'IPE_1_4_-1': 'Assembleia|AGESP|   TODAS'
                                                                                    #  'IPE_1_2_-1': 'Assembleia|AGO|   TODAS'
                                                                                    #  'IPE_1_3_-1': 'Assembleia|AGO/E|   TODAS'
                                                                                    #  'IPE_3_-1_-1': 'Aviso aos Acionistas|TODOS|   TODAS'
                                                                                    #  'IPE_9_-1_-1': 'Calendário de Eventos Corporativos|TODOS|   TODAS'
                                                                                    #  'IPE_91_-1_-1': 'Carta Anual de Governança Corporativa|TODOS|   TODAS'                                         

# search_result = search_result[
#     (search_result['categoria']=="DFP - Demonstrações Financeiras Padronizadas") |
#     (search_result['categoria']=="ITR - Informações Trimestrais")]# |
    # (search_result['categoria']=="FCA - Formulário Cadastral") |
    # (search_result['categoria']=="FRE - Formulário de Referência") |
    # (search_result['categoria']=="IAN - Informações Anuais")]



search_result = search_result[pd.to_numeric(search_result['numero_seq_documento'], errors='coerce').notnull()]


print(search_result)


agora = datetime.datetime.now().strftime("%Y%m%d__%H_%M_%S")

# Receber o valor do documento do usuário
valor_documento = input("Digite o número do documento que está buscando: ")

# Converter para inteiro
valor_documento = int(valor_documento)

# Filtrar os resultados da consulta externa da CVM pelo valor do documento
search_result = search_result[search_result['numero_seq_documento'] == valor_documento]

# Abrir a URL da consulta externa da CVM
response = urllib3.urlopen(f"{valor_documento}")
# Ler o conteúdo HTML da página
html = response.read()

# Criar um objeto BeautifulSoup para analisar o HTML
soup = BeautifulSoup(html, "html.parser")

# Encontrar todas as tags <a> que contêm os links dos documentos
links = soup.find_all("a", href=True)

# Inicializar a variável url como vazia
url = ""

# Percorrer os links e comparar o valor do documento com o número_seq_documento que está na URL do link
for link in links:
    # Obter o valor do atributo href do link
    href = link.get("href")
    # Verificar se o href contém o valor do documento
    if str(valor_documento) in href:
        # Atribuir o href à variável url
        url = href
        # Sair do laço
        break

# Verificar se a variável url não está vazia
if url != "":
    # Lendo a tabela da página web
    tabela = pd.read_html(url, header=0)[0]

    # Carregando o arquivo Excel com o template pré-formatado
    book = load_workbook("template.xlsx")

    # Definindo o writer para escrever em um novo arquivo
    writer = pd.ExcelWriter(f"arquivo_editado_{valor_documento}.xlsx", engine="openpyxl") # Incluir o valor do documento no nome do arquivo

    # Incluindo a formatação no writer
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Escrevendo os dados em uma planilha específica, na posição desejada
    tabela.to_excel(writer, "Sheet1", startrow=1, startcol=1, header=False, index=False)

# Criar uma nova planilha em branco
wb = openpyxl.Workbook()
planilha = wb.active
planilha.title = 'Resultado do REPORT'

# Adicionar os nomes das colunas à planilha
planilha.append(['Conta', 'Descrição', 'Valor', 'currency_unit', 'cod_cvm'])

reports_list = [
    'Balanço Patrimonial Ativo',
    'Balanço Patrimonial Passivo',
    'Demonstração do Resultado',
    'Demonstração do Resultado Abrangente',
    'Demonstração do Fluxo de Caixa',
    'Demonstração das Mutações do Patrimônio Líquido',
    'Demonstração de Valor Adicionado'] # Se None retorna todos os demonstrativos disponíveis.

# Percorrer os resultados e adicionar os dados à planilha
for index, row in search_result.iterrows():
    empresa = f"{row['cod_cvm']} - {cvm_codes[row['cod_cvm']]}"
    print("*" * 20, empresa, "*" * 20)
    # Usar um bloco try-except para capturar possíveis exceções
    try:
        # Especificar os nomes dos relatórios que quer obter
        reports = cvm_httpclient.get_report(row["numero_seq_documento"], row["codigo_tipo_instituicao"], reports_list=reports_list)
        for report in reports:
            reports[report]["cod_cvm"] = row["cod_cvm"]
            print(reports[report].head())
            # Converter o dataframe em um array numpy
            data = reports[report].to_numpy()
            # Adicionar os dados à planilha
            for row in data:
                planilha.append(row.tolist())
    except ValueError as e:
        # Mostrar uma mensagem de aviso se o valor do documento não existir
        print(f"Valor do documento inválido: {e}")
    except IndexError as f:
        # Mostrar uma mensagem de aviso se houver algum problema na conexão
        print(f"Erro na requisição: {f}")
    except KeyError as g:
        print(f'Erro de Valor: {g}')

# Salvando o arquivo
writer.save()
